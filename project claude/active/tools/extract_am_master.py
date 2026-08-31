# -*- coding: utf-8 -*-
"""
Ekstrak master data AM Checksheet dari file .xlsx warisan menjadi tabel
ternormalisasi yang siap di-seed ke Google Sheets (backend webapp AM/PM Tracker).

Menangani kekacauan file warisan:
  - Layout kolom berbeda antara file Adult dan Baby Pants -> kolom dideteksi
    dari header, bukan hardcode.
  - Header "Kegiatan" berperan sebagai Standar, bukan Tindakan.
  - Sheet Packer di file Baby bernama generik Sheet2..Sheet6 -> stasiun
    disimpulkan dari isi kolom PIC.
  - Sheet "Weekly Monthly (Week NN)" adalah arsip hasil audit, bukan master,
    sehingga dipisah ke am_history.csv.

Output:
  seed/am_tasks.csv    master tasklist AM (siap jadi tab MST_AM_TASK)
  seed/am_history.csv  arsip hasil audit mingguan warisan
  seed/am_master.json  gabungan + metadata dokumen
"""
import os
import re
import json
import csv
import hashlib
from collections import Counter

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # .../active
SEED = os.path.join(BASE, "seed")

SOURCES = [
    {"file": "Checksheet AM Adult Rev.00.xlsx",
     "line": "AHP", "rev": "00", "machines": ["AHP1"]},
    {"file": "Checksheet AM Baby Pants Rev.4 (BHP 1 - BHP 3).xlsx",
     "line": "BHP", "rev": "04", "machines": ["BHP1", "BHP2", "BHP3"]},
]


def find_source(name):
    """
    Cari file .xlsx sumber dengan menelusuri folder ke atas.

    File warisan tersimpan di root workspace, sementara script ini berada di
    'project claude/active/tools'. Penelusuran ke atas membuat script tetap
    jalan meski struktur folder digeser lagi.
    """
    here = os.path.dirname(os.path.abspath(__file__))
    for _ in range(6):
        candidate = os.path.join(here, name)
        if os.path.isfile(candidate):
            return candidate
        parent = os.path.dirname(here)
        if parent == here:
            break
        here = parent
    raise FileNotFoundError(
        "File sumber tidak ditemukan di folder mana pun di atas tools/: " + name)

FREQ_MARKERS = {
    "SHIFTLY": "SHIFTLY", "DAILY": "DAILY", "WEEKLY": "WEEKLY",
    "MONTHLY": "MONTHLY", "2 WEEKLY": "BIWEEKLY",
    "3 MONTHLY": "QUARTERLY", "6 MONTHLY": "SEMIANNUAL", "YEARLY": "YEARLY",
}

# "kegiatan" dipakai sebagai kriteria penerimaan di sheet warisan -> standard
HEADER_ALIASES = {
    "no.": "no", "no": "no",
    "nama part": "part",
    "tindakan": "action",
    "standar kebersihan": "standard",
    "kegiatan": "standard",
    "pic": "pic",
    "check": "check",
}

SKIP_RE = re.compile(r"^(vis\.|full vis|keterangan perubahan|list door|copy of)", re.I)
ARCHIVE_RE = re.compile(r"^weekly monthly \(week\s*(\d+)\)", re.I)
STATION_RE = re.compile(r"^(OP\s*\d+|Packer\s*\d*|Palleting\s*\d*)", re.I)


def norm(v):
    return "" if v is None else re.sub(r"\s+", " ", str(v)).strip()


def find_header(ws, max_scan=12):
    """Cari baris header, petakan peran kolom -> index kolom (1-based)."""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        cells = {}
        for c in range(1, min(ws.max_column, 60) + 1):
            role = HEADER_ALIASES.get(norm(ws.cell(r, c).value).lower())
            if role:
                cells.setdefault(role, c)
        if "no" in cells and "part" in cells:
            return r, cells
    return None, None


def sheet_title(ws):
    for r in range(1, 6):
        for c in range(1, 14):
            m = re.search(r"(CHECKLIST[^\n]*)", norm(ws.cell(r, c).value), re.I)
            if m:
                return norm(m.group(1))
    return ""


def doc_meta(ws):
    labels = {
        "no. dokumen": "doc_no", "revisi": "doc_rev", "tgl berlaku": "doc_effective",
        "dibuat oleh": "doc_author", "disetujui oleh": "doc_approver",
    }
    meta = {}
    for r in range(1, 7):
        for c in range(1, min(ws.max_column, 60) + 1):
            key = norm(ws.cell(r, c).value).lower().rstrip(":")
            if key in labels:
                for cc in range(c + 1, min(c + 4, ws.max_column + 1)):
                    val = norm(ws.cell(r, cc).value).lstrip(":").strip()
                    if val:
                        meta[labels[key]] = val
                        break
    return meta


def canon_station(raw):
    """Normalkan 'OP 1' / 'Packer 2 (Wanita)' / 'Palleting 1' ke bentuk kanonik."""
    m = STATION_RE.match(norm(raw))
    if not m:
        return ""
    return re.sub(r"\s+", "", m.group(1).upper())


def parse_stations(raw):
    """
    Pecah label PIC gabungan menjadi daftar stasiun kanonik.

    'OP 1,2,3/Packer'  -> ['OP1', 'OP2', 'OP3', 'PACKER']
    'OP 1, OP 2, OP3'  -> ['OP1', 'OP2', 'OP3']
    'Packer 2 (Pria)'  -> ['PACKER2']
    """
    text = norm(raw).upper()
    if not text:
        return []
    out, prefix = [], None
    for token in re.split(r"[,/&]| DAN ", text):
        token = token.strip()
        if not token:
            continue
        m = re.match(r"^(OP|PACKER|PALLETING|PALLETI\w*)\s*(\d*)", token)
        if m:
            prefix = "PALLETING" if m.group(1).startswith("PALLETI") else m.group(1)
            out.append(prefix + m.group(2))
            continue
        # angka telanjang mewarisi prefix sebelumnya, mis. "3" dalam "OP 1,2,3"
        m = re.match(r"^(\d+)", token)
        if m and prefix:
            out.append(prefix + m.group(1))
    seen = []
    for s in out:
        if s not in seen:
            seen.append(s)
    return seen


def is_task_no(v):
    return bool(re.fullmatch(r"\d+(\.\d+)?", norm(v)))


def slug(*parts):
    return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:10].upper()


def read_sheet(ws, hrow, cols):
    """Yield (freq, seq, part, action, standard, pic, check) untuk tiap baris task."""
    freq = None
    for r in range(hrow + 1, ws.max_row + 1):
        a = norm(ws.cell(r, cols["no"]).value)
        if a.upper() in FREQ_MARKERS:
            freq = FREQ_MARKERS[a.upper()]
            continue
        if not is_task_no(a):
            continue

        def col(role):
            return norm(ws.cell(r, cols[role]).value) if role in cols else ""

        part, action = col("part"), col("action")
        std, pic, chk = col("standard"), col("pic"), col("check")
        if not (part or action or std):
            continue
        if not part:  # baris lanjutan mewarisi Nama Part dari baris di atasnya
            for back in range(r - 1, hrow, -1):
                prev = norm(ws.cell(back, cols["part"]).value)
                if prev:
                    part = prev
                    break
        yield freq or "SHIFTLY", int(float(a)), part, action, std, pic, chk


def extract():
    tasks, history, docs = [], [], []
    for src in SOURCES:
        wb = load_workbook(find_source(src["file"]), data_only=True)
        for ws in wb.worksheets:
            if SKIP_RE.search(ws.title):
                continue
            hrow, cols = find_header(ws)
            if not hrow:
                continue
            parsed = list(read_sheet(ws, hrow, cols))
            if not parsed:
                continue

            archive = ARCHIVE_RE.match(ws.title)
            if archive:
                week = int(archive.group(1))
                for freq, seq, part, _action, std, pic, chk in parsed:
                    history.append({
                        "line": src["line"], "week": week, "frequency": freq,
                        "seq": seq, "part_name": part, "standard": std, "pic": pic,
                        "result": {"V": "OK", "X": "NG"}.get(chk.upper(), chk),
                    })
                docs.append({"kind": "archive", "line": src["line"],
                             "sheet": ws.title, "week": week,
                             "task_count": len(parsed)})
                continue

            # Stasiun kanonik = PIC yang paling sering muncul, fallback ke nama sheet
            pics = [canon_station(p[5]) for p in parsed if canon_station(p[5])]
            if pics:
                station = Counter(pics).most_common(1)[0][0]
            else:
                m = re.search(r"(OP\s*\d+|PACKER\s*\d*|PALLETI\w*)", ws.title, re.I)
                station = canon_station(m.group(1)) if m else "UNASSIGNED"

            meta = doc_meta(ws)
            hidden = ws.sheet_state != "visible"
            for freq, seq, part, action, std, pic, _chk in parsed:
                # Stasiun diambil dari PIC baris itu sendiri; fallback ke stasiun sheet
                stations = parse_stations(pic) or [station]
                tasks.append({
                    "task_id": slug(src["line"], station, freq, str(seq), part, action),
                    "line": src["line"],
                    "machines": ";".join(src["machines"]),
                    "station": station,
                    "stations": ";".join(stations),
                    "frequency": freq,
                    "seq": seq,
                    "part_name": part,
                    "action": action,
                    "standard": std,
                    "pic_label": pic or station,
                    "doc_no": meta.get("doc_no", ""),
                    "doc_rev": src["rev"],
                    "doc_effective": meta.get("doc_effective", ""),
                    "source_sheet": ws.title,
                    # sheet tersembunyi di file warisan = versi lama yang sudah digantikan
                    "active": "FALSE" if hidden else "TRUE",
                })
            docs.append({"kind": "master", "line": src["line"], "sheet": ws.title,
                         "station": station, "title": sheet_title(ws),
                         "hidden": hidden, "task_count": len(parsed), **meta})
        wb.close()
    return tasks, history, docs


def write_csv(path, rows, fields):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def main():
    os.makedirs(SEED, exist_ok=True)
    tasks, history, docs = extract()

    write_csv(os.path.join(SEED, "am_tasks.csv"), tasks,
              ["task_id", "line", "machines", "station", "stations", "frequency",
               "seq", "part_name", "action", "standard", "pic_label", "doc_no",
               "doc_rev", "doc_effective", "source_sheet", "active"])
    write_csv(os.path.join(SEED, "am_history.csv"), history,
              ["line", "week", "frequency", "seq", "part_name", "standard",
               "pic", "result"])
    with open(os.path.join(SEED, "am_master.json"), "w", encoding="utf-8") as f:
        json.dump({"documents": docs, "tasks": tasks, "history": history},
                  f, ensure_ascii=False, indent=1)

    print("master task  :", len(tasks))
    print("arsip audit  :", len(history), "baris")
    print("per lini     :", dict(Counter(t["line"] for t in tasks)))
    print("per frekuensi:", dict(Counter(t["frequency"] for t in tasks)))
    print("aktif/nonaktif:", dict(Counter(t["active"] for t in tasks)))
    beban = Counter()
    for t in tasks:
        if t["active"] == "TRUE":
            for s in t["stations"].split(";"):
                beban[s] += 1
    print("beban/stasiun:", dict(sorted(beban.items())))
    print("hasil arsip  :", dict(Counter(h["result"] for h in history)))
    print()
    print("master per sheet:")
    for d in docs:
        if d["kind"] == "master":
            print("  {:4} {:11} {:32} {:>3} task {}".format(
                d["line"], d["station"], d["sheet"][:32], d["task_count"],
                "(hidden)" if d["hidden"] else ""))


if __name__ == "__main__":
    main()
